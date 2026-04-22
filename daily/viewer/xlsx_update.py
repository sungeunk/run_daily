#!/usr/bin/env python3
"""Append one column of daily results to the shared master xlsx.

Workflow
--------
The team keeps a master xlsx on SharePoint/OneDrive. Each row identifies a
benchmark configuration (model / precision / in_token / out_token / exec).
Each column is one daily run, stamped with commit id, workweek, datetime.

This module:

1. Opens the master xlsx (expected to be locally synced by OneDrive).
2. Reads the key columns (model, precision, in, out, exec) to learn row order
   — so there is no ``FIXED_ROW_ORDER`` constant to maintain anymore.
3. Appends a new column at the right edge, with the three header cells
   (commit, workweek, datetime) and one value per matching row.
4. Saves in place. OneDrive / SharePoint takes over the sync.

The xlsx layout is discovered, not prescribed:

* ``--sheet`` picks which sheet to write to (default: first sheet).
* ``--key-cols`` is a 1-based tuple of columns that hold
  (model, precision, in, out, exec). Default: ``A,B,C,D,E``.
* ``--header-rows`` is the number of rows above the first data row. Default
  is 3 — commit/ww/date — matching the old viewer's paste block.
* Rows whose key matches nothing in the summary stay blank.

The master xlsx is *not* edited in place unless the caller explicitly
passes the path. Without ``--xlsx-update``, run.py does nothing here.
"""

from __future__ import annotations

import argparse
import logging
from dataclasses import dataclass
from pathlib import Path
from typing import Any


log = logging.getLogger(__name__)


@dataclass
class XlsxTarget:
    path: Path
    sheet: str | None = None     # None → first visible sheet
    key_cols: tuple[int, ...] = (1, 2, 3, 4, 5)  # 1-based (A=1 … E=5)
    header_rows: int = 3         # commit / ww / datetime
    data_start_row: int | None = None  # defaults to header_rows + 1


def _normalise(v: Any) -> Any:
    """Normalise a key cell value to something hashable / comparable.

    * Excel ints come back as int already; floats that are actually ints
      (e.g. 32.0) get collapsed to int so keys compare with perf_rows output.
    * Strings are stripped — whitespace in the master xlsx shouldn't block
      matches.
    """
    if v is None:
        return 0
    if isinstance(v, float) and v.is_integer():
        return int(v)
    if isinstance(v, str):
        s = v.strip()
        # Treat a blank cell the same as 0 for numeric key columns.
        if s == '':
            return ''
        return s
    return v


def _read_row_keys(ws, key_cols, data_start_row, last_row) -> list[tuple]:
    """Return list of row-key tuples for rows ``data_start_row..last_row``.

    A tuple with no non-blank values ends the data block — we stop there so
    trailing empty rows don't accidentally get populated.
    """
    keys = []
    for r in range(data_start_row, last_row + 1):
        row_vals = tuple(_normalise(ws.cell(row=r, column=c).value)
                         for c in key_cols)
        # Blank row = end of data.
        if all(v in ('', 0, None) for v in row_vals):
            break
        keys.append(row_vals)
    return keys


def update_master_xlsx(
    target: XlsxTarget,
    lookup: dict[tuple, float],
    headers: tuple[str, str, str],
    *,
    value_format: str = '0.00',
) -> tuple[int, int]:
    """Append one column of values to ``target``.

    Returns ``(matched, total)`` — how many rows were populated vs how many
    data rows the sheet has, so the caller can report coverage.
    """
    # Imported lazily so ``python run.py --help`` doesn't require openpyxl.
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter

    wb = load_workbook(target.path)
    ws = wb[target.sheet] if target.sheet else wb[wb.sheetnames[0]]

    data_start_row = target.data_start_row or (target.header_rows + 1)
    last_row = ws.max_row
    keys = _read_row_keys(ws, target.key_cols, data_start_row, last_row)

    # New column goes to the first truly empty column to the right.
    new_col = ws.max_column + 1
    new_col_letter = get_column_letter(new_col)
    log.info('xlsx: writing column %s (rows %d..%d)',
             new_col_letter, data_start_row, data_start_row + len(keys) - 1)

    # Header block (commit / ww / datetime by default).
    for i, h in enumerate(headers):
        ws.cell(row=i + 1, column=new_col, value=h)

    matched = 0
    for offset, key in enumerate(keys):
        # Key is (model, precision, in, out, exec) — same tuple perf_rows emits.
        val = lookup.get(key)
        if val is None:
            # Try a token-category fallback: if the xlsx key has 'short'/'long'
            # strings, match on those; otherwise leave blank.
            continue
        cell = ws.cell(row=data_start_row + offset, column=new_col, value=val)
        cell.number_format = value_format
        matched += 1

    wb.save(target.path)
    return matched, len(keys)


# ---------------------------------------------------------------------------
# CLI (for manual updates outside run.py)
# ---------------------------------------------------------------------------

def _parse_cli() -> argparse.Namespace:
    p = argparse.ArgumentParser(
        description='Append one daily-run column to the shared master xlsx.',
        formatter_class=argparse.ArgumentDefaultsHelpFormatter,
    )
    p.add_argument('--summary', type=Path, required=True,
                   help='Path to daily.<stamp>.summary.json')
    p.add_argument('--xlsx', type=Path, required=True,
                   help='Path to the locally-synced master xlsx')
    p.add_argument('--sheet', default=None)
    p.add_argument('--key-cols', default='1,2,3,4,5',
                   help='1-based column indices holding (model,precision,in,out,exec)')
    p.add_argument('--header-rows', type=int, default=3)
    return p.parse_args()


def _main() -> int:
    import json
    args = _parse_cli()
    logging.basicConfig(level=logging.INFO,
                        format='%(asctime)s [%(levelname).1s] %(message)s')

    # Import siblings without adding sys.path gymnastics when invoked as a
    # module from repo root: ``python -m daily.viewer.xlsx_update ...``.
    from .perf_rows import flatten, as_lookup

    summary = json.loads(args.summary.read_text(encoding='utf-8'))
    lookup = as_lookup(flatten(summary))

    meta = summary.get('meta', {})
    headers = (
        meta.get('ov_version', ''),
        meta.get('workweek', ''),
        meta.get('stamp', ''),
    )

    target = XlsxTarget(
        path=args.xlsx,
        sheet=args.sheet,
        key_cols=tuple(int(x) for x in args.key_cols.split(',')),
        header_rows=args.header_rows,
    )
    matched, total = update_master_xlsx(target, lookup, headers)
    log.info('xlsx update done: %d/%d rows matched', matched, total)
    return 0


if __name__ == '__main__':
    raise SystemExit(_main())
